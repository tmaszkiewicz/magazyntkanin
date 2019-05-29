# Barcode
from reportlab.graphics.barcode import qr, eanbc
from reportlab.graphics.shapes import Drawing
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.graphics import renderPDF
# Praca z PDF
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger

# Praca z linia komend
import os
from openpyxl import load_workbook
from subprocess import call
import re
import shutil

# CSV
import unicodecsv

# Import baz danych
from .models import Tkanina, \
    Rolka, \
    Dziennik, \
    WpisySzwalnia, \
    WpisyMagazyn, \
    Log, \
    ErrorLog

# Praca z datami
from datetime import datetime
from django.http import HttpResponse

def Tworzenie_CSV(path=os.path.join('.', 'magazyntkanin', 'pliki', 'euro'),data_dostawy=''):
    return_string = ""
    for scan in os.scandir(path):
        if re.search(r'.xlsx$', scan.name) and scan.is_file():
            try:
                file_name_csv = scan.path[:len(scan.path)-4] + 'csv'
                call(["xlsx2csv", scan.path, file_name_csv])
                return_string += Import_euro(file_name_csv,data_dostawy)+"\n"
                print(return_string)
                shutil.move(scan.path, os.path.join(path, 'archive'))
                os.remove(file_name_csv)
            except Exception as e:
                return e
    return return_string

def Import_euro(plik, data_dostawy):
    try:
        licznik = 0
        with open(plik, 'rb') as f:
            csvdata = unicodecsv.reader(f, delimiter=',', encoding='utf-8')
            header = next(csvdata)
            data = []
            for i in csvdata:
                if not i == header:
                    data.append(i)
                   
        for row in data:
            if len(row) > 1:
                dlugosc = float(row[2])
                nazwa_tkaniny = row[1].strip().upper()
                if Tkanina.objects.filter(nazwa=nazwa_tkaniny).exists():
                    tkanina = Tkanina.objects.get(nazwa=nazwa_tkaniny)
                    index, created = Rolka.objects.get_or_create(
                        tkanina=tkanina,
                        nr_rolki=row[4],
                        dlugosc=dlugosc,
                        dlugosc_poczatkowa=dlugosc,
                        nr_zamowienia=row[0],
                        barcode=row[6],
                        lot=row[3],
                        data_dostawy=data_dostawy)
                    print(data_dostawy)
                    if created:
                        licznik += 1
                else:
                    None
                    #raise Exception
    except Exception as e:
        print(e)
    print(licznik)
    return str("Dodano {0} rolek").format(str(licznik))

def importBazyTkanin(plik):
    with open(plik, 'rb') as f:
        csvdata = unicodecsv.reader(f, delimiter=',', encoding='utf-8')
        ilosc = 0
        for i in csvdata:
            index, created = Tkanina.objects.get_or_create(
                index_sap=i[0], nazwa=i[1].strip())
            if created:
                ilosc += 1
        return str("Utworzono {0} rekordow").format(str(ilosc))

def pdf_view(dziennik):
    path = os.path.join('.', 'dzienniki', 'krojownia')
    with open(path + "/" + dziennik + '.pdf', 'rb') as pdf:
        response = HttpResponse(pdf.read(), mimetype='application/pdf')
        response['Content-Disposition'] = 'inline;filename=dziennik.pdf'
        return response
    pdf.closed

def createBarCodes(barcode_value, temp_path):
    """
    Create barcode examples and embed in a PDF
    """
    c = canvas.Canvas(temp_path + ".bar", pagesize=landscape(A4))
    print(temp_path)
    # draw a QR code
    qr_code = qr.QrCodeWidget(str(barcode_value))
    bounds = qr_code.getBounds()

    width = bounds[2] - bounds[0]
    height = bounds[3] - bounds[1]
    d = Drawing(75, 75, transform=[75. / width, 0, 0, 75. / height, 0, 0])
    d.add(qr_code)
    # renderPDF.draw(d, c, 26 * cm, 18.5 * cm)
    renderPDF.draw(d, c, 0.2 * cm, 26 * cm)
    c.save()
    return True

def tkaniny_A4_barcode(tkanina, sap_index, file_name="tmp/tabela.pdf", wszystkie_=False):
    from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.graphics.barcode import createBarcodeDrawing

    c = SimpleDocTemplate(file_name, pagesize=A4, topMargin=0.5*cm, bottomMargin=0.5*cm)
    styleSheet = getSampleStyleSheet()
    w = styleSheet['Normal']
    w.fontSize = 24
    w.leading = w.fontSize * 1.2

    elements = []
    data = []

    if wszystkie_:
        tkaniny = Tkanina.objects.all().order_by('pk')
        licznik = 0
        wszystkie = []
        for tkanina in tkaniny:
            licznik += 1
            P = Paragraph("<para align=center ><b>{0}</b></para>".format(tkanina.nazwa), w)
            Barcode = createBarcodeDrawing('Code128', value=tkanina.index_sap, height=1.2*cm, width=8*cm, fontSize=8, humanReadable = True)
            Barcode.height = 1.5*cm
            if licznik % 3 == 0:
                data.append([P, Barcode])
                wszystkie.append(data)
                data = []
                continue
            data.append([P, Barcode])
        t=Table(wszystkie, rowHeights = 4.7 * cm, colWidths = 6.5 * cm)
    else:
        P = Paragraph("<para align=center ><b>{0}</b></para>".format(tkanina), w)
        Barcode = createBarcodeDrawing('Code128', value=sap_index, height=1.2*cm, width=8*cm, fontSize=8, humanReadable = True)
        Barcode.height = 1.5*cm
        for i in range(6):
            data.append([[P, Barcode],[P, Barcode],[P, Barcode]])
        t=Table(data, rowHeights = 4.7 * cm, colWidths = 6.5 * cm)
    t.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER'),
                        ('TOPPADDING', (0,0), (-1,-1), cm),
                        ('BOTTOMPADDING', (0,0), (-1,-1), cm),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                    ]))
    elements.append(t)
    c.build(elements)
    return True

def generuj_rap_inwentury():

    from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate
    #from reportlab.platypus import Table, TableStyle, Paragraph, 
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate, PageBreak, NextPageTemplate
    c = canvas.Canvas('tmp/inv.pdf', pagesize=A4)
    width, height = A4
    top = height - 100
    left=50
    rowsForPage = 300
    cols = [ x*100+left for x in range(5)]
    print(cols)

    # dodanie czcionek
    pdfmetrics.registerFont(TTFont('AlegreyaSC', 'magazyntkanin/static/fonts/Alegreya_SC/AlegreyaSC-Bold.ttf'))
    # paragraf
    stylesheet=getSampleStyleSheet()
    styleN = stylesheet['Normal']
    styleN.fontSize = 17

    p = Paragraph(u'<para align="center"><b>KOREKTY INWENTARYZACYJNE</b></para>', styleN)
    w,h = p.wrap(width, 200)
    p.drawOn(c, 5, height-70)

    korekty = Log.objects.filter(typ="INWENTURA")
    c.setFont("Times-Bold", 8)
    row = 100
    c.drawString(cols[0], top, "rolka_id")
    c.drawString(cols[1], top, "index_tkaniny")
    c.drawString(cols[2], top, "przed korekta")
    c.drawString(cols[3], top, "po korekcie")
    c.drawString(cols[4], top, "data/godzina")
    row+=10
    page = 0


    for korekta in korekty:
    #for korekta in range(100):
        row +=10
     #   c.drawString(cols[0], top-(row % rowsForPage),"aaa")
        c.drawString(cols[0], height-(row % rowsForPage), str(korekta.rolka_id))
        c.drawString(cols[1], height-(row % rowsForPage), str(korekta.index_tkaniny))
        c.drawString(cols[2], height-(row % rowsForPage), str(korekta.dlugosc_rolki))
        c.drawString(cols[3], height-(row % rowsForPage), str(korekta.dlugosc_elementu))
        dt = datetime.fromtimestamp(1544778947).isoformat()
        c.drawString(cols[4], height-(row % rowsForPage), dt)
        

        c.drawString(width/2-4, 10, "str. "+str(page))
        if(row % rowsForPage == 0): 
            c.showPage()
            c.setFont("Times-Bold", 8)
            page+=1

    #c.setFont("Times-Bold", 8)
    #c.drawString(100, 10, "ME05")
    #c.drawString(width-125, 10, "Dane w [mm]")

    # zapis do pliku
    c.showPage()
    c.save()
def modulo_no_zero(row,rowsForPage,top,height):
   
    if row % rowsForPage!=0:
        return height-(row % rowsForPage)-top
    else:
        return height - rowsForPage -top

def generuj_raport_xls(inw):
    import datetime
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "INW"
    row=1
    for i in inw:
        cell = "A" + str(row)
        try:
            ws[cell] = i['nazwa_tkaniny']
        except:
            None
        cell = "B" + str(row)
        try:
            ws[cell] = i['index_tkaniny']
        except:
            None
        cell = "C" + str(row)
        try:
            ws[cell] = i['rolka_id']
        except:
            None
        cell = "D" + str(row)
        try:
            ws[cell] = i['dlugosc_rolki']
        except:
            None
        cell = "E" + str(row)
        try:
            ws[cell] = i['dlugosc_elementu']
        except:
            None
        
        cell = "F" + str(row)
        try:
            ws[cell] = i['dlPerIndeks']
        except:
            None
        cell = "G" + str(row)
        try:
            ws[cell] = i['dlPerIndeks_inw']
        except:
            None
        cell = "H" + str(row)
        try:
            ws[cell] = i['isPrinted']
        except:
            None
        cell = "I" + str(row)
        try:
            ws[cell] = i['typ']
        except:
            None
        cell = "J" + str(row)
        try:
            ws[cell] = i['timestamp']
        except:
            None
        row+=1


    wb.save('tmp/inw-'+datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")+'.xlsx')
    #wb.save('inw.xlsx')


    return HttpResponse("OK")
def generuj_xls_porownanie_sap(inw):
    import datetime
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "INW"
    row=1
    for i in inw:
        if  i['isPrinted']:
            cell = "A" + str(row)
            try:
                ws[cell] = i['nazwa_tkaniny']
            except:
                None
            cell = "B" + str(row)
            try:
                ws[cell] = i['index_tkaniny']
            except:
                None
            #cell = "C" + str(row)
            #try:
            #    ws[cell] = i['rolka_id']
            #except:
            #    None
            #cell = "D" + str(row)
            #try:
            #    ws[cell] = i['dlugosc_rolki']
            #except:
            #    None
            #cell = "E" + str(row)
            #try:
            #    ws[cell] = i['dlugosc_elementu']
            #except:
            #    None
            
            cell = "C" + str(row)
            try:
                ws[cell] = i['dlPerIndeks']
            except:
                None
            cell = "D" + str(row)
            try:
                ws[cell] = i['dlPerIndeks_inw']
            except:
                None
            #cell = "H" + str(row)
            #try:
            #    ws[cell] = i['isPrinted']
            #except:
            #    None
            #cell = "I" + str(row)
            #try:
            #    ws[cell] = i['typ']
            #except:
            #    None
            #cell = "J" + str(row)
            #try:
            #    ws[cell] = i['timestamp']
            #except:
            #    None
            row+=1
    #wb.save('tmp/inw-'+datetime.datetime.now().strftime("%Y-%m-%d-%H-%M")+'.xlsx')
    try:
        os.remove('tmp/inwsrc.xlsx')
    except:
        print("Nie ma pliku do usuniecia")
    wb.save('tmp/inwsrc.xlsx')
    os.chmod('tmp/inwsrc.xlsx',0o777)

    #wb.save('inw.xlsx')

def generuj_raport_inwentury2(inw):

    from reportlab.platypus import Table, TableStyle, Paragraph, SimpleDocTemplate
    #from reportlab.platypus import Table, TableStyle, Paragraph, 
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate, PageBreak, NextPageTemplate
    c = canvas.Canvas('tmp/inv.pdf', pagesize=A4)
    width, height = A4
    
    top = 70
    left=50
    rowsForPage = 700
    cols = [ x*80+left for x in range(6)]
    print(cols)

    # dodanie czcionek
    pdfmetrics.registerFont(TTFont('AlegreyaSC', 'magazyntkanin/static/fonts/Alegreya_SC/AlegreyaSC-Bold.ttf'))
    # paragraf
    stylesheet=getSampleStyleSheet()
    styleN = stylesheet['Normal']
    styleN.fontSize = 17

    p = Paragraph(u'<para align="center"><b>RAPORT INWENTARYZACJI</b></para>', styleN)
    w,h = p.wrap(width, 200)
    p.drawOn(c, 5, height -50)

    #korekty = Log.objects.filter(typ="INWENTURA")

    c.setFont("Times-Bold", 8)
    row = 10
    page = 1

    c.drawString(width/2-4, 10, "str. "+str(page))

    for i in inw:
        if i['isPrinted']:
            None
            #row+=10
            #c.drawString(cols[0], height-(row % rowsForPage), str(i['nazwa_tkaniny']))
            #c.drawString(cols[1], height-(row % rowsForPage), str(i['index_tkaniny']))
            #c.drawString(cols[2], height-(row % rowsForPage), str(i['dlPerIndeks']))
            #c.drawString(cols[3], height-(row % rowsForPage), str(i['dlPerIndeks_inw']))
        else:
            None
        if i['typ'] == 'INWENTURA':
            c.setFont("Times-Bold", 8)
            #color = red
        else:
            c.setFont("Times-Roman", 8)
                
        if i['isPrinted']:
            row+=20
            # row % rowsForPage - zeruje sie na modulo i ostatni wiersz ze strony trafia na poczatek
            # trzeba chyba obliczac height-(row % rowsForPage i jesli row % rowsForPag 0 to max na stronie
            
            c.drawString(cols[0], modulo_no_zero(row,rowsForPage,top,height), str(i['nazwa_tkaniny']))
            c.drawString(cols[1], modulo_no_zero(row,rowsForPage,top,height), str(i['index_tkaniny']))
            c.drawString(cols[2], modulo_no_zero(row,rowsForPage,top,height), str(i['dlPerIndeks']))
            c.drawString(cols[3], modulo_no_zero(row,rowsForPage,top,height), str(i['dlPerIndeks_inw']))
            row+10
            c.drawString(cols[0], modulo_no_zero(row,rowsForPage,top,height), "__________________________________________________________________________________")
            row+=10
            c.drawString(cols[0], modulo_no_zero(row,rowsForPage,top,height), "Nazwa_tkaniny")
            c.drawString(cols[1], modulo_no_zero(row,rowsForPage,top,height), "index_tkaniny")
            c.drawString(cols[2], modulo_no_zero(row,rowsForPage,top,height), "rolka_id")
            c.drawString(cols[3], modulo_no_zero(row,rowsForPage,top,height), "przed korekta")
            c.drawString(cols[4], modulo_no_zero(row,rowsForPage,top,height), "po korekcie")
            c.drawString(cols[5], modulo_no_zero(row,rowsForPage,top,height), "data/godzina")
            row+=10
            c.drawString(cols[0], modulo_no_zero(row,rowsForPage,top,height), str(i['nazwa_tkaniny']))
            c.drawString(cols[1], modulo_no_zero(row,rowsForPage,top,height), str(i['index_tkaniny']))
            c.drawString(cols[2], modulo_no_zero(row,rowsForPage,top,height), str(i['rolka_id']))
            c.drawString(cols[3], modulo_no_zero(row,rowsForPage,top,height), str(i['dlugosc_rolki']))
            c.drawString(cols[4], modulo_no_zero(row,rowsForPage,top,height), str(i['dlugosc_elementu']))

            #POPRAW DATY!
            #PODZEL STRONY!
            
            #dt = datetime.fromtimestamp(int(str(i['timestamp']))).isoformat()
            try:
                c.drawString(cols[5], height-(row % rowsForPage)-top, str(i['timestamp'].strftime("%Y-%m-%d %H:%M")))
            except:
                None
            
        else:
            row+=10
            #hpos = modulo_no_zero(row,rowsForPage,top)
            c.drawString(cols[2], modulo_no_zero(row,rowsForPage,top,height), str(i['rolka_id']))
            c.drawString(cols[3], modulo_no_zero(row,rowsForPage,top,height), str(i['dlugosc_rolki']))
            c.drawString(cols[4], modulo_no_zero(row,rowsForPage,top,height), str(i['dlugosc_elementu']))

            #dt = datetime.fromtimestamp(1544778947).isoformat()
            #dt = datetime.fromtimestamp(int(str(i['timestamp']))).isoformat()
            #c.drawString(cols[5], modulo_no_zero(row,rowsForPage,top,height) , dt)
            try:
                c.drawString(cols[5], modulo_no_zero(row,rowsForPage,top,height) , str(i['timestamp'].strftime("%Y-%m-%d %H:%M")))
            except:
                None

        if row % rowsForPage >= 690: 
            c.showPage()
            c.setFont("Times-Bold", 8)
            page+=1
            c.drawString(width/2-4, 10, "str. "+str(page))
            row=10








    #for korekta in korekty:
    #    row +=10
    #    c.drawString(cols[0], height-(row % rowsForPage), str(korekta.rolka_id))
    #    c.drawString(cols[1], height-(row % rowsForPage), str(korekta.index_tkaniny))
    #    c.drawString(cols[2], height-(row % rowsForPage), str(korekta.dlugosc_rolki))
    #    c.drawString(cols[3], height-(row % rowsForPage), str(korekta.dlugosc_elementu))
    #    dt = datetime.fromtimestamp(1544778947).isoformat()
    #    c.drawString(cols[4], height-(row % rowsForPage), dt)
        

    #    c.drawString(width/2-4, 10, "str. "+str(page))
    #    if(row % rowsForPage == 0): 
    #        c.showPage()
    #        c.setFont("Times-Bold", 8)
    #        page+=1

    #c.setFont("Times-Bold", 8)
    #c.drawString(100, 10, "ME05")
    #c.drawString(width-125, 10, "Dane w [mm]")

    # zapis do pliku
    c.showPage()
    c.save()

def generuj_obiegowke_v1(
    id="",
    index="",
    nazwa_tkaniny="",
    lot="",
    rolka="",
    data_zamowienia="",
    szerokosc="",
    dlugosc="",
    qr_draw=False,
    full=False
    ):
    # -*- coding: utf-8 -*-
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib import colors

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A5
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF
    wiersz=0
    strnr=1
    c = canvas.Canvas('tmp/obiegowka.pdf', pagesize=A5)
    width, height = A5
    # dodanie czcionek
    pdfmetrics.registerFont(TTFont('AlegreyaSC', 'magazyntkanin/static/fonts/Alegreya_SC/AlegreyaSC-Bold.ttf'))

    index = index
    nazwa_sap = nazwa_tkaniny
    lot = lot
    rolka = rolka
    data_zamowienia = str(data_zamowienia)
    # if isinstance(data_zamowienia, datetime):
    #     data_zamowienia = data_zamowienia.strftime("%d/%m/%Y")
    if data_zamowienia == "None":
        data_zamowienia = ""
    szerokosc = szerokosc
    dlugosc = dlugosc

    #QR
    if qr_draw == True:
        qr_code = qr.QrCodeWidget(str(id))
        d = Drawing(45, 45)
        d.add(qr_code)
        d.wrap(45,45)
        d.drawOn(c, width-90, height-90)
        c.setFont("Times-Bold", 10)
        c.drawString(width-55, height-90, str(id))

    # paragraf
    stylesheet=getSampleStyleSheet()
    styleN = stylesheet['Normal']
    styleN.fontSize = 17
    p = Paragraph(u'<para align="center"><b>Karta obiegowa tkaniny w Janipolu</b></para>', styleN)
    w,h = p.wrap(width-100, 200)
    p.drawOn(c, 5, height-20)

    # -------------------------------
    data = []
    data.append(['Index', index])
    data.append(['Nazwa tkaniny', nazwa_sap])
    t=Table(data, [100, 220])
    t.setStyle(TableStyle([('FONT', (0,0), (0,-1), 'AlegreyaSC', 9),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('GRID', (0,0), (-1,-1), 1, colors.black),
                        # ('BACKGROUND', (0, 0), (0, 1), colors.lightgrey),
                    ]))
    w,h = t.wrap(width, height)
    t.drawOn(c, 10, (height-75))

    # -------------------------------
    data = []
    data.append(['Lot', lot])
    data.append(['Nr Rolki', rolka])
    c.setFont("Times-Bold", 10)
    c.drawString(width-150, height-150, "DATA: " + data_zamowienia)
    t=Table(data, [110, 70])
    t.setStyle(TableStyle([('FONT', (0,0), (0,-1), 'AlegreyaSC', 9),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                        # ('BACKGROUND', (0, 0), (0, 1), colors.lightgrey),
                    ]))
    w,h = t.wrap(width, height)
    t.drawOn(c, 10, (height-150))

    # -------------------------------
    data = []
    data.append(['Ogólna szerokość (mm)', szerokosc])
    data.append(['Krajka lewa', ""])
    data.append(['Krajka prawa', ""])
    data.append(['Szer. do krojenia', ""])
    t=Table(data, [110, 70])
    t.setStyle(TableStyle([('FONT', (0,0), (0,-1), 'AlegreyaSC', 9),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                        # ('BACKGROUND', (0, 0), (0, 3), colors.lightgrey),
                    ]))
    w,h = t.wrap(width, height)
    t.drawOn(c, 10, (height-225))

    # -------------------------------
    data = []
    data.append(['Długość (mb)', dlugosc])
    t=Table(data, [70, 50])
    t.setStyle(TableStyle([('FONT', (0,0), (0,-1), 'AlegreyaSC', 9),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                        # ('BACKGROUND', (0, 0), (0, 1), colors.lightgrey),
                    ]))
    w,h = t.wrap(width, height)
    t.drawOn(c, 290, (height-250))

    # -------------------------------
    data = []
    data.append(['Data', 'Nazwa układu', 'C/R', 'Podpis', 'Zużyto', 'Zwrócono'])
    wiersz =12

    if full:
        #wpisy = Log.objects.filter(rolka_id=id, typ__startswith="FGK")
        wpisy = Log.objects.filter(rolka_id=id).order_by('-timestamp')

        for each in wpisy:
            
            try:
                nrfgk=each.nr_fgk
                typfgk=each.typ
                if not each.nr_fgk:
                    nrfgk=each.typ

            except:
                nrfgk=""
            try:
                dlrolki=each.dlugosc_rolki
                if not each.dlugosc_rolki:
                    dlrolki=0            
            except:
                dlrolki = 0
            try:
                dlelementu=each.dlugosc_elementu
                if not each.dlugosc_elementu:
                    dlelementu=0            
            except:
                dlelementu = 0
        
            print(nrfgk)
            
            if typfgk=="EDYCJA":
                data.append([each.timestamp.strftime('%d-%m-%Y'), nrfgk, rolka, '', dlrolki, round(dlelementu, 2)])
            elif typfgk in ("WYDANIE_MAG_WZORNIKI", "FGK_poza","FGK","FGK_laczone","FGK_laczone_poza"):
                data.append([each.timestamp.strftime('%d-%m-%Y'), nrfgk, rolka, '', dlelementu, round(dlrolki, 2)])

            else:
                data.append([each.timestamp.strftime('%d-%m-%Y'), nrfgk, rolka, '', dlelementu, round(dlrolki-dlelementu, 2)])
                
        for i in range(13-wpisy.count()):
            data.append(['', '', '', '', '', ''])
    else:
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
    wiersz = 0 
    dat = []
    for i in data:
        wiersz+=1
        if strnr==1:
            rowPerPage=14
        else:
            rowPerPage=24
        dat.append(i)
        if wiersz % rowPerPage==0:  
            wiersz = 0
            strnr+=1
            try:
                t=Table(dat, [50, 130, 50, 70, 50, 50],rowHeights=23)
                t.setStyle(TableStyle([('FONT', (0,0), (-1,0), 'AlegreyaSC', 9),
                                    ('FONT', (0,1), (-1,-1), 'Times-Roman', 9),
                                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                                    ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                                ]))
                w,h = t.wrap(width, 300)
                t.drawOn(c, 10, 20)
            except:
                None

            # -------------------------------
            c.setFont("Times-Bold", 8)
            c.drawString(100, 10, "ME05")
            c.drawString(width-125, 10, "Dane w [mm]"+str(wiersz))
            dat = []

            # zapis do pliku
            c.showPage()
    #print(dat)
    
    while wiersz % rowPerPage!=0:
        wiersz+=1
        dat.append(['', '', '', '', '', ''])        


    try:
        t=Table(dat, [50, 130, 50, 70, 50, 50],rowHeights=23)
        t.setStyle(TableStyle([('FONT', (0,0), (-1,0), 'AlegreyaSC', 9),
                            ('FONT', (0,1), (-1,-1), 'Times-Roman', 9),
                            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                            ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                        ]))
        w,h = t.wrap(width, 300)
        t.drawOn(c, 10, 20)
    except:
        None

    # -------------------------------
    c.setFont("Times-Bold", 8)
    c.drawString(100, 10, "ME05")
    c.drawString(width-125, 10, "Dane w [mm]"+str(wiersz))
    dat = []

    #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    #t=Table(dat, [50, 130, 50, 70, 50, 50],rowHeights=23)
    #t.setStyle(TableStyle([('FONT', (0,0), (-1,0), 'AlegreyaSC', 9),
    #                    ('FONT', (0,1), (-1,-1), 'Times-Roman', 9),
    #                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    #                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    #                    ('GRID', (0,0), (-1,-1), 0.25, colors.black),
    #                ]))
    #w,h = t.wrap(width, 300)
    #t.drawOn(c, 10, 20)

    # -------------------------------
    #c.setFont("Times-Bold", 8)
    #c.drawString(100, 10, "ME05")
    #c.drawString(width-125, 10, "Dane w [mm]"+str(wiersz))

    # zapis do pliku
    c.showPage()

    c.save()
def generuj_obigowke(
    id="",
    index="",
    nazwa_tkaniny="",
    lot="",
    rolka="",
    data_zamowienia="",
    szerokosc="",
    dlugosc="",
    qr_draw=False,
    full=False
    ):
    # -*- coding: utf-8 -*-
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib import colors

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A5
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.graphics.barcode import qr
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics import renderPDF

    c = canvas.Canvas('tmp/obiegowka.pdf', pagesize=A5)
    width, height = A5
    # dodanie czcionek
    pdfmetrics.registerFont(TTFont('AlegreyaSC', 'magazyntkanin/static/fonts/Alegreya_SC/AlegreyaSC-Bold.ttf'))

    index = index
    nazwa_sap = nazwa_tkaniny
    lot = lot
    rolka = rolka
    data_zamowienia = str(data_zamowienia)
    # if isinstance(data_zamowienia, datetime):
    #     data_zamowienia = data_zamowienia.strftime("%d/%m/%Y")
    if data_zamowienia == "None":
        data_zamowienia = ""
    szerokosc = szerokosc
    dlugosc = dlugosc

    #QR
    if qr_draw == True:
        qr_code = qr.QrCodeWidget(str(id))
        d = Drawing(45, 45)
        d.add(qr_code)
        d.wrap(45,45)
        d.drawOn(c, width-90, height-90)
        c.setFont("Times-Bold", 10)
        c.drawString(width-55, height-90, str(id))

    # paragraf
    stylesheet=getSampleStyleSheet()
    styleN = stylesheet['Normal']
    styleN.fontSize = 17
    p = Paragraph(u'<para align="center"><b>Karta obiegowa tkaniny w Janipolu</b></para>', styleN)
    w,h = p.wrap(width-100, 200)
    p.drawOn(c, 5, height-20)

    # -------------------------------
    data = []
    data.append(['Index', index])
    data.append(['Nazwa tkaniny', nazwa_sap])
    t=Table(data, [100, 220])
    t.setStyle(TableStyle([('FONT', (0,0), (0,-1), 'AlegreyaSC', 9),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('GRID', (0,0), (-1,-1), 1, colors.black),
                        # ('BACKGROUND', (0, 0), (0, 1), colors.lightgrey),
                    ]))
    w,h = t.wrap(width, height)
    t.drawOn(c, 10, (height-75))

    # -------------------------------
    data = []
    data.append(['Lot', lot])
    data.append(['Nr Rolki', rolka])
    c.setFont("Times-Bold", 10)
    c.drawString(width-150, height-150, "DATA: " + data_zamowienia)
    t=Table(data, [110, 70])
    t.setStyle(TableStyle([('FONT', (0,0), (0,-1), 'AlegreyaSC', 9),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                        # ('BACKGROUND', (0, 0), (0, 1), colors.lightgrey),
                    ]))
    w,h = t.wrap(width, height)
    t.drawOn(c, 10, (height-150))

    # -------------------------------
    data = []
    data.append(['Ogólna szerokość (mm)', szerokosc])
    data.append(['Krajka lewa', ""])
    data.append(['Krajka prawa', ""])
    data.append(['Szer. do krojenia', ""])
    t=Table(data, [110, 70])
    t.setStyle(TableStyle([('FONT', (0,0), (0,-1), 'AlegreyaSC', 9),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                        # ('BACKGROUND', (0, 0), (0, 3), colors.lightgrey),
                    ]))
    w,h = t.wrap(width, height)
    t.drawOn(c, 10, (height-225))

    # -------------------------------
    data = []
    data.append(['Długość (mb)', dlugosc])
    t=Table(data, [70, 50])
    t.setStyle(TableStyle([('FONT', (0,0), (0,-1), 'AlegreyaSC', 9),
                        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                        # ('BACKGROUND', (0, 0), (0, 1), colors.lightgrey),
                    ]))
    w,h = t.wrap(width, height)
    t.drawOn(c, 290, (height-250))

    # -------------------------------
    data = []
    data.append(['Data', 'Nazwa układu', 'C/R', 'Podpis', 'Zużyto', 'Zwrócono'])
    if full:
        wpisy = Log.objects.filter(rolka_id=id, typ__startswith="FGK")
        for each in wpisy:
            data.append([each.timestamp.strftime('%d-%m-%Y'), each.nr_fgk, rolka, '', each.dlugosc_elementu, round(each.dlugosc_rolki-each.dlugosc_elementu, 2)])
        for i in range(13-wpisy.count()):
            data.append(['', '', '', '', '', ''])
    else:
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
        data.append(['', '', '', '', '', ''])
    t=Table(data, [50, 130, 50, 70, 50, 50],rowHeights=23)
    t.setStyle(TableStyle([('FONT', (0,0), (-1,0), 'AlegreyaSC', 9),
                        ('FONT', (0,1), (-1,-1), 'Times-Roman', 9),
                        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),

                        ('GRID', (0,0), (-1,-1), 0.25, colors.black),
                    ]))
    w,h = t.wrap(width, 300)
    t.drawOn(c, 10, 20)

    # -------------------------------
    c.setFont("Times-Bold", 8)
    c.drawString(100, 10, "ME05")
    c.drawString(width-125, 10, "Dane w [mm]")

    # zapis do pliku
    c.showPage()
    c.save()

def roznice_w_plikach():
    a = []
    b = []
    for each in os.walk(os.path.join('.', 'dzienniki', 'krojownia')):
        for file in each[2]:
            if not file == 'Thumbs.db':
                a.append(file)
    for each in os.walk(os.path.join('.', 'dzienniki', 'magazyn')):
        for file in each[2]:
            if not file == 'Thumbs.db':
                b.append(file)
    magazyn = set(b).difference(a)
    krojownia = set(a).difference(b)
    return krojownia, magazyn

def pdf_merger(pdf_list):
    merger = PdfFileMerger()
    root_path = os.path.join('.', 'dzienniki')
    krojownia_path = os.path.join(root_path, 'krojownia')
    magazyn_path = os.path.join(root_path, 'magazyn')
    for pdf in pdf_list:
        try:
            merger.append(
                open(os.path.join(krojownia_path, str(pdf) + '.pdf'), 'rb'))
        except Exception as e:
            continue
    for pdf in pdf_list:
        try:
            merger.append(
                open(os.path.join(magazyn_path, str(pdf) + '.pdf'), 'rb'))
        except Exception as e:
            continue
    with open('tmp/output_marge.pdf', 'wb') as file_out:
        merger.write(file_out)

def Tworzenie_kodu_na_dzienniku(nr_zlecenia, dziennik_path, barcode_path='temp_barcode.pdf'):
    output = PdfFileWriter()
    if createBarCodes(nr_zlecenia, dziennik_path):
        a_pdf = PdfFileReader(open(dziennik_path, 'rb'))
        b_pdf = PdfFileReader(open(dziennik_path + ".bar", 'rb'))
        barcode = b_pdf.getPage(0)
        for i in range(a_pdf.getNumPages()):
            page = a_pdf.getPage(i)
            page.mergePage(barcode)
            output.addPage(page)
        with open(dziennik_path + "_", 'wb') as f:
            output.write(f)
        os.remove(dziennik_path + ".bar")
        return True

def generuj_etykiete_tkaniny(nazwa_tkaniny, barcode=[], L='', R='', data_dostawy='', M='', szerokosc=''):
    c = canvas.Canvas("tmp/etykieta.pdf", pagesize=(7.8 * cm, 5.4 * cm))
    ilosc = len(barcode)
    for i, etykieta in enumerate(range(ilosc)):
        if len(nazwa_tkaniny) < 15:
            c.setFont('Courier-Bold', size=24)
        elif len(nazwa_tkaniny) > 15 and len(nazwa_tkaniny) < 20:
            c.setFont('Courier-Bold', size=18)
        else:
            c.setFont('Courier-Bold', size=14)
        c.drawCentredString(3.9 * cm, 4.5 * cm, nazwa_tkaniny)
        qr_code = qr.QrCodeWidget(barcode[i])
        bounds = qr_code.getBounds()
        width = bounds[2] - bounds[0]
        height = bounds[3] - bounds[1]
        d = Drawing(60, 60, transform=[60. / width, 0, 0, 60. / height, 0, 0])
        d.add(qr_code)
        renderPDF.draw(d, c, 2.8 * cm, 2 * cm)
        c.setFont('Courier-Bold', size=10)
        c.drawString(3.5 * cm, 3.9 * cm, barcode[i])
        c.setFont('Courier-Bold', size=18)
        c.drawString(0.2 * cm, 3.2 * cm, 'L:')
        c.setFont('Courier-Bold', size=14)
        c.drawString(1.1 * cm, 3.2 * cm, L)
        c.drawString(1.1 * cm, 3.2 * cm, "_____")
        c.setFont('Courier-Bold', size=18)
        c.drawString(0.2 * cm, 1.7 * cm, 'R:')
        c.setFont('Courier-Bold', size=14)
        c.drawString(1.1 * cm, 1.7 * cm, R)
        c.drawString(1.1 * cm, 1.7 * cm, "_____")
        c.setFont('Courier-Bold', size=18)
        c.drawString(4.8 * cm, 3.2 * cm, 'M:')
        c.setFont('Courier-Bold', size=14)
        if not M == '':
            c.drawString(5.6 * cm, 3.2 * cm, str(float(M)) + ' m')
        c.drawString(5.6 * cm, 3.2 * cm, "_____")
        c.setFont('Courier-Bold', size=18)
        c.drawString(4.4 * cm, 1.7 * cm, 'SZ:')
        c.setFont('Courier-Bold', size=14)
        if not szerokosc == '':
            c.drawString(5.6 * cm, 1.7 * cm, str(szerokosc) + ' mm')
        c.drawString(5.6 * cm, 1.7 * cm, "_____")
        c.setFont('Courier-Bold', size=18)
        c.drawString(1.87 * cm, 0.9 * cm, 'Data:')
        c.setFont('Courier-Bold', size=14)
        if isinstance(data_dostawy, datetime):
            c.drawString(3.9 * cm, 0.9 * cm, data_dostawy.strftime("%d/%m/%Y"))
        else:
            c.drawString(3.9 * cm, 0.9 * cm, data_dostawy)
        c.showPage()
    c.save()

def generuj_etykiete_tkaniny_podwojna(nazwa_tkaniny, barcode=[], L='', R='', data_dostawy='', M='', szerokosc=''):
    pegewidth=8.8
    shift = 4.8  #4.4  #5.4
    pageheight = 5.4+shift # 10.2
    qrsize = 1.0

    l_shift = 0.3 #0.7 
    c = canvas.Canvas("tmp/etykieta_podwojna.pdf", pagesize=(pegewidth * cm, pageheight * cm))
    ilosc = len(barcode)
    #DZIELIMY NA NAZWE I KOLOR
    nazwa_tkaniny=nazwa_tkaniny.split(sep=" ")
    for i, etykieta in enumerate(range(ilosc)):
        print(len(nazwa_tkaniny[0]))

        if len(nazwa_tkaniny[0]) <14  and len(nazwa_tkaniny[0]) < 14:
            c.setFont('Courier-Bold', size=28)

        elif (len(nazwa_tkaniny[0])>=14 or len(nazwa_tkaniny[1])>=14)  and len(nazwa_tkaniny[0]) < 20 and len(nazwa_tkaniny[1]) < 20:
            c.setFont('Courier-Bold', size=20)
        else:
            c.setFont('Courier-Bold', size=18)

        #NAZWA      
        c.drawCentredString((pegewidth/2) * cm, (4.5+shift-0.5) * cm, nazwa_tkaniny[0])
        #KOLOR
        c.drawCentredString((pegewidth/2) * cm, (4.5+shift-1.5) * cm, nazwa_tkaniny[1])

        qr_code = qr.QrCodeWidget(barcode[i])
        bounds = qr_code.getBounds()
        width = bounds[2] - bounds[0]
        height = bounds[3] - bounds[1]
        d = Drawing(60, 60, transform=[60. / width, 0, 0, 60. / height, 0, 0])
        d.add(qr_code)
        renderPDF.draw(d, c, (pegewidth/2 - 2/2-0.5) * cm, (2+shift-2.0) * cm)

        c.setFont('Courier-Bold', size=10)
        c.drawCentredString((pegewidth/2+0.1-0.5) * cm, (3.9+shift-2.0) * cm, barcode[i])

        c.setFont('Courier-Bold', size=20)
        c.drawString((0.2+l_shift) * cm, (3.2+shift-1.5) * cm, 'L:')
        c.setFont('Courier-Bold', size=14)
        c.drawString((1.1+l_shift) * cm, (3.2+shift-1.5) * cm, L)
        c.drawString((1.1+l_shift) * cm, (3.2+shift-1.5) * cm, "_____")
        c.setFont('Courier-Bold', size=20)
        c.drawString((0.2+l_shift) * cm, (1.7+shift-1.5) * cm, 'R:')
        c.setFont('Courier-Bold', size=14)
        c.drawString((1.1+l_shift) * cm, (1.7+shift-1.5) * cm, R)
        c.drawString((1.1+l_shift) * cm, (1.7+shift-1.5) * cm, "_____")
        c.setFont('Courier-Bold', size=20)
        c.drawString((5.0+l_shift) * cm, (3.2+shift-1.5) * cm, 'M:')
        c.setFont('Courier-Bold', size=14)
        if not M == '':
            c.drawString((5.8+l_shift) * cm, (3.2+shift-1.5) * cm, str(float(M)) + ' m')
        c.drawString((5.8+l_shift) * cm, (3.2+shift-1.5) * cm, "______")
        c.setFont('Courier-Bold', size=20)
        c.drawString((4.6+l_shift) * cm, (1.7+shift-1.5) * cm, 'SZ:')
        c.setFont('Courier-Bold', size=14)
        if not szerokosc == '':
            c.drawString((5.8+l_shift) * cm, (1.7+shift-1.5) * cm, str(szerokosc) + 'mm')
        c.drawString((5.8+l_shift) * cm, (1.7+shift-1.5) * cm, "______")
        c.setFont('Courier-Bold', size=20)
                     #1.87
        c.drawString((1.2+l_shift) * cm, (0.9+shift-2.0) * cm, 'Data:')
        c.setFont('Courier-Bold', size=18)
        if isinstance(data_dostawy, datetime):
            c.drawString((3.4+l_shift) * cm, (0.9+shift-2.0) * cm, data_dostawy.strftime("%d/%m/%Y"))
        else:
            c.drawString((3.4+l_shift) * cm, (0.9+shift-2.0) * cm, data_dostawy)

        #dodanie kodu ponizej  Tomek 19.11.2018
        

        qr_code = qr.QrCodeWidget(barcode[i])
        bounds = qr_code.getBounds()
        width = bounds[2] - bounds[0]
        height = bounds[3] - bounds[1]
        d = Drawing(60, 60, transform=[60. / width, 0, 0, 60. / height, 0, 0])
        d.add(qr_code)
        #renderPDF.draw(d, c, (2.8+l_shift) * cm, 0 * cm)
        c.setFont('Courier-Bold', size=10)
        c.drawString((0.5+l_shift) * cm, 2.3 * cm, barcode[i])
        c.drawString((5.7+l_shift) * cm, 2.3 * cm, barcode[i])
        #c.drawString((4.9+l_shift) * cm, 1,5 * cm, barcode[i])
        renderPDF.draw(d, c, (0.0+l_shift) * cm, 0.4 * cm)
        renderPDF.draw(d, c, (5.2+l_shift) * cm, 0.4 * cm)

        #eanbc_code = eanbc.Ean8BarcodeWidget(barcode[i])
        #bounds = eanbc_code.getBounds()
        #width = bounds[2] - bounds[0]
        #height = bounds[3] - bounds[1]
        #d = Drawing(60, 60, transform=[60. / width, 0, 0, 60. / height, 0, 0])
        #d.add(eanbc_code)
        #renderPDF.draw(d, c, (2.8+l_shift) * cm, 0 * cm)

        c.showPage()
    c.save()

def test_ulotka_():
    return call(['lp -djan-pr-mm01 -o portrait -o fit-to-page -o media=A5 tmp/wzornik_tmp.pdf'], shell=True)

def Bezposredni_wydruk(Rolka):
    rolka = Rolka
    L = rolka.lot
    if not L:
        L = ''
    R = rolka.nr_rolki
    if not R:
        R = ''
    M = rolka.dlugosc
    if not M:
        M = ''
    try:
        data_dostawy = rolka.data_dostawy.strftime("%d/%m/%Y")
    except Exception as e:
        data_dostawy = ''
    nazwa_tkaniny = rolka.tkanina.nazwa
    barcode = [str(rolka.pk)]
    generuj_etykiete_tkaniny_podwojna(nazwa_tkaniny, barcode, str(
        L), str(R), data_dostawy, str(M))
    call(['/etc/init.d/cups start'], shell=True)
    #call(['lp tmp/etykieta.pdf'], shell=True)

    #call(['lp tmp/etykieta_podwojna.pdf'], shell=True) #### 06.05.2019
    call(['lp -o Resolution=300dpi -o PageSize=w144h216 tmp/etykieta_podwojna.pdf'], shell=True) #06.05.2019
    call(['lp -djan-pr-mm01 -o portrait -o fit-to-page -o media=A5 tmp/obiegowka.pdf'], shell=True)

def odczytaj_csv(plik_path=os.path.join('.', 'dzienniki')):
    przekazywana_tablica = []
    with open(plik_path, 'rb') as f:
        zawartosc = unicodecsv.reader(f, delimiter='\t', encoding='ISO-8859-1')
        for each in zawartosc:
            przekazywana_tablica.append(each)
    return przekazywana_tablica

def odczytaj_txt(plik_path=os.path.join('.', 'dzienniki')):
    przekazywana_tablica = []
    with open(plik_path, 'r', encoding='ISO-8859-1') as f:
        lines = f.readlines()
        for each in lines:
            przekazywana_tablica.append(each)
    return przekazywana_tablica

def rozpoznawanie_dziennika(zawartosc_pliku):
    flag = "M"
    for each in zawartosc_pliku:
        if re.search(r'Informacja', each):
            flag = "K"                    
    return flag

def dziennik_krojowania(zawartosc_pliku):
    # --- Odczytanie Nr Dziennika
    nr_dziennika = zawartosc_pliku[2][1]
    # --- Odczytanie daty
    re_data_zaplanowana = r'Informacja.+(\d{1,2}[,./]\d{1,2}[,./]\d{4})'
    data_zaplanowana = re.search(re_data_zaplanowana, zawartosc_pliku[5][0])
    data_zaplanowana = data_zaplanowana.group(1)
    # --- Wyizolowanie wpisów
    tablica_wpisow = []
    # --- Zmienne poczatkowe
    wpisy = {}
    lp=0
    ta=None
    # --- Petla przeszukujaca wpisy
    for linia in zawartosc_pliku:
        try:
            #Pominiecie poczatkowych linii z tekstem i naglowka tabeli
            if re.search(r'[a-zA-Z]', linia[0]) or not linia[2]:
                continue
            # Jezeli OCR odczytal tylko TA
            if re.search(r'\d{7}', linia[1]):
                lp += 1
                ta = linia[1]
                wpisy = {
                        'lp': lp,
                        'ta': ta,
                        'index_sap': linia[3],
                        'tura': linia[4],
                        'ilosc': linia[7]
                        }
                print("Main", wpisy)
                tablica_wpisow.append(wpisy)
                continue
            elif not linia[1] and linia[3]:
                # Sprawdzenie czy nie jest to kontynuacja wpisu
                wpisy = {
                    'lp': lp,
                    'ta': ta,
                    'index_sap': linia[3],
                    'tura': linia[4],
                    'ilosc': linia[7]
                }
                if not wpisy in tablica_wpisow:
                    print("Dodatkowy", wpisy)
                    tablica_wpisow.append(wpisy)
                continue
        except IndexError:
            continue
    return nr_dziennika, data_zaplanowana, tablica_wpisow

def dziennik_krojowania_re(zawartosc_pliku):
    # # --- Odczytanie daty
    # re_data_zaplanowana = r'Informacja.+(\d+[,./]\d+[,./]\d+)'
    re_data_zaplanowana = r'Informacja.+\s(\d+[,.\/]\d+[,.\/]\d+)'

    # # --- Wyizolowanie wpisów
    tablica_wpisow = []
    # --- Zmienne poczatkowe
    wpisy = {}
    lp=0
    ta=None
    # --- Petla przeszukujaca wpisy
    re_ta = r'\d{7}'
    re_sap_index_tura = r'(\d{5})\s+(\d{4,5})'
    re_ilosc = r'\s(\d{3})\s'
    for numer, linia in enumerate(zawartosc_pliku):
        if re.search(r'^[a-zA-Z]', linia):
            # --- Odczytanie Nr Dziennika
            nr_dziennika = re.search(r'Zlec. zbiorcze\s+(\d+)', linia)
            if nr_dziennika:
                nr_dziennika_ = nr_dziennika.group(1)                                       
            data_zaplanowana = re.search(re_data_zaplanowana, linia)
            if data_zaplanowana:
                data_zaplanowana_ = data_zaplanowana.group(1)
            continue
        else:                
            ta_search = re.search(re_ta, linia)
            sap_index_tura_search = re.search(re_sap_index_tura, linia)
            ilosc_search = re.search(re_ilosc, linia)            
            if ta_search:
                lp = lp + 1
                try:
                    ta = ta_search.group(0)
                    sap_index = sap_index_tura_search.group(1)
                    tura = sap_index_tura_search.group(2)
                except AttributeError:
                    print(numer, linia)
                    print(f'TA: {ta} / Index_sap: {sap_index} / Tura: {tura}')
                    int(linia)
                    continue
                ilosc = int(ilosc_search.group(1))
                wpisy = {
                    'lp': lp,
                    'ta': ta,
                    'index_sap': sap_index,
                    'tura': tura,
                    'ilosc': ilosc
                    }
                tablica_wpisow.append(wpisy)
                continue
            elif sap_index_tura_search:
                sap_index = sap_index_tura_search.group(1)
                tura = sap_index_tura_search.group(2)
                ilosc = int(ilosc_search.group(1))
                wpisy = {
                    'lp': lp,
                    'ta': ta,
                    'index_sap': sap_index,
                    'tura': tura,
                    'ilosc': ilosc
                    }                
                if not wpisy in tablica_wpisow:
                    if ilosc > 0:
                        wpisy_check = {
                            'lp': lp,
                            'ta': ta,
                            'index_sap': sap_index,
                            'tura': tura,
                            'ilosc': 1
                            }
                        if not wpisy_check in tablica_wpisow:
                            tablica_wpisow.append(wpisy)
    if not nr_dziennika_ or not data_zaplanowana_:
        raise ValueError('Niepoprana data lub nr_dziennika')
    return nr_dziennika_, data_zaplanowana_, tablica_wpisow

def dziennik_krojowania_re_test(zawartosc_pliku):
    import re
    # # --- Odczytanie daty    
    # re_data_zaplanowana = r'Informacja.+\s(\d+[,.\/]\d+[,.\/]\d+)'
    # re_data_zaplanowana = r'Informacja.+\s[\d*]+[-,]*(\d+[,.\/]\d+[,.\/]\d+)'
    re_data_zaplanowana = r'Data wydruk\s+(\d+[,.\/]\d+[,.\/]\d+)'
    

    # # --- Wyizolowanie wpisów
    tablica_wpisow = []
    # --- Zmienne poczatkowe
    wpisy = {}
    lp=0
    ta=None
    ilosc=1
    # --- Petla przeszukujaca wpisy
    re_ta = r'\|(\d{7})\|'
    re_sap_index_tura = r'\|(\d{5})\s+\|(\d{4,5})\s*\|'
    re_lp = r'^\|\s*(\d+)\s\|'
    re_shop = r'SHOP'
    for numer, linia in enumerate(zawartosc_pliku):
        if re.search(r'^\s+[a-zA-Z]', linia):
            # --- Odczytanie Nr Dziennika
            nr_dziennika = re.search(r'Zlec. zbiorcze\s+(\d+)', linia)
            if nr_dziennika:
                nr_dziennika_ = nr_dziennika.group(1)                                       
            data_zaplanowana = re.search(re_data_zaplanowana, linia)
            if data_zaplanowana:
                data_zaplanowana_ = data_zaplanowana.group(1)
            continue
        else:
            if re.search(r'^\|-', linia) or re.search(r'^-', linia):
                continue
            if re.search(r'^\|', linia):
                if re.search(re_shop, linia):
                    re_sap_index_tura = r'\|(\d{5})\s+\|SHOP\s*\|'
                    tura = '00000'
                ta_search = re.search(re_ta, linia)
                sap_index_tura_search = re.search(re_sap_index_tura, linia)
                lp_search = re.search(re_lp, linia)
                if ta_search:
                    lp = lp_search.group(1)
                    ta = ta_search.group(1)
                
                    sap_index = sap_index_tura_search.group(1)
                    if not re.search(re_shop, linia):
                        tura = sap_index_tura_search.group(2)
                    wpisy = {
                        'lp': lp,
                        'ta': ta,
                        'index_sap': sap_index,
                        'tura': tura,
                        'ilosc': 1
                        }
                    
                    tablica_wpisow.append(wpisy)
                    continue
                elif sap_index_tura_search:
                    sap_index = sap_index_tura_search.group(1)
                    try:
                        tura = sap_index_tura_search.group(2)
                    except:
                        None
                    print ("dziennik krojownia te test")
                    wpisy = {
                        'lp': lp,
                        'ta': ta,
                        'index_sap': sap_index,
                        'tura': tura,
                        'ilosc': 1
                        }                
                    if not wpisy in tablica_wpisow:
                        if ilosc > 0:
                            wpisy_check = {
                                'lp': lp,
                                'ta': ta,
                                'index_sap': sap_index,
                                'tura': tura,
                                'ilosc': 1
                                }
                            if not wpisy_check in tablica_wpisow:
                                tablica_wpisow.append(wpisy)
    print(nr_dziennika_,data_zaplanowana_,tablica_wpisow)
    if not nr_dziennika_ or not data_zaplanowana_:
        raise ValueError('Niepoprana data lub nr_dziennika')
    return nr_dziennika_, data_zaplanowana_, tablica_wpisow

def dziennik_magazyn(zawartosc_pliku):
    # --- Wyszukanie nr dziennika
    zlecenie = re.search(r'\d+', zawartosc_pliku[1][0])
    zlecenie = zlecenie.group(0)
    # --- zmienne
    tablica_wpisow = []
    # --- Petla wyszukujaca wpisy
    for each in zawartosc_pliku:
        try:
            if re.search(r'\d', each[2]): # --- jezeli tablica jest dluzsza niz 3 i zawiera liczbe w 3 kolumnie to jest to wpis
                pozycja = {
                        'index_sap': int(each[0].strip()),
                        'dlugosc': float(each[2].replace(',', '.'))
                        }
                tablica_wpisow.append(pozycja)
        except IndexError:
            # --- w przypadku kiedy tablica jest mniejsza niz 3 elementy, sprawdz nastepny
            continue
    return zlecenie, tablica_wpisow

def dziennik_magazyn_re(zawartosc_pliku):
    # --- Wyszukanie nr dziennika
    re_zlecenie = r'(\d{7})'    

    # --- zmienne
    tablica_wpisow = []
    kombinacje = r'^\|(\d+).+\|\s+(\d+,\d+)'
    # --- Petla wyszukujaca wpisy
    for each in zawartosc_pliku:
        linia_kombinacji = re.search(kombinacje, each)
        _zlecenie = re.search(re_zlecenie, each)
        if _zlecenie:
            zlecenie = _zlecenie.group(0)
        if linia_kombinacji: # --- jezeli tablica jest dluzsza niz 3 i zawiera liczbe w 3 kolumnie to jest to wpis
            pozycja = {
                    'index_sap': int(linia_kombinacji.group(1).strip()),
                    'dlugosc': float(linia_kombinacji.group(2).replace(',', '.'))
                    }
            tablica_wpisow.append(pozycja)
        continue
    return zlecenie, tablica_wpisow

def kopiuj_dane_dziennika():
    error_array = []
    for scan in os.scandir(os.path.join('.', 'dzienniki', 'input')):
        if scan.is_file:            
            sap_print_name = re.search(r'C\d+' ,scan.name)
            if sap_print_name:
                nazwa_pliku = sap_print_name.group(0)
                poszukiwany_plik = nazwa_pliku + ".TXT"
                nazwa_pdf = nazwa_pliku + ".pdf"
                path_dziennikow_txt = os.path.join('.', 'dzienniki', 'input', 'TXT')            
                try:
                    shutil.copy(os.path.join(path_dziennikow_txt, poszukiwany_plik), os.path.join('.', 'dzienniki', 'backup'))
                    shutil.copy(scan.path, os.path.join('.', 'dzienniki', 'backup'))                    
                    shutil.move(os.path.join(path_dziennikow_txt, poszukiwany_plik), os.path.join('.', 'dzienniki'))
                    shutil.move(scan.path, os.path.join('.', 'dzienniki', nazwa_pdf))                    
                except Exception as e:
                    print(e)
                    error_array.append(scan.name)
            else:
                continue
    return error_array

def Zarzadzenie_dziennikami(ocr_txt_path=os.path.join('.', 'dzienniki/')):
    krojowania = []
    error = []
    magazyn = []    
    brak_danych = kopiuj_dane_dziennika()
    for scan in os.scandir(ocr_txt_path):        
        if re.search(r'.TXT$', scan.name) and scan.is_file():
            data = []
            csv_z_pliku = odczytaj_csv(plik_path=scan.path)
            txt_z_pliku = odczytaj_txt(plik_path=scan.path)
            file_pdf_name = scan.path[:len(scan.path) - 4] + '.pdf'
            dzial = rozpoznawanie_dziennika(txt_z_pliku)
            if dzial == "M":
                try:
                    nr_zlecenia, dane = dziennik_magazyn_re(txt_z_pliku)
                    if Tworzenie_kodu_na_dzienniku(nr_zlecenia, file_pdf_name):
                        Zapisywanie_danych(dane, 'M', nr_zlecenia)
                        magazyn.append((file_pdf_name, nr_zlecenia))
                        os.remove(scan.path)
                        os.remove(file_pdf_name)
                        shutil.move(
                            file_pdf_name + '_',
                            os.path.join('.', 'dzienniki', 'magazyn', str(nr_zlecenia) + '.pdf')
                            )
                except Exception as e:
                    print(f'{file_pdf_name} - {e}')
            elif dzial == "K":
                try:
                    nr_zlecenia, data_zlecenia, dane = dziennik_krojowania_re_test(txt_z_pliku)
                    
                    if Tworzenie_kodu_na_dzienniku(nr_zlecenia, file_pdf_name):
                        Zapisywanie_danych(dane, 'K', nr_zlecenia, data_zlecenia)
                        krojowania.append((file_pdf_name, nr_zlecenia))
                        os.remove(scan.path)
                        os.remove(file_pdf_name)
                        shutil.move(
                            file_pdf_name + '_',
                            os.path.join('.', 'dzienniki', 'krojownia', str(nr_zlecenia) + '.pdf')
                            )                        
                except Exception as e:
                    print(f'{file_pdf_name} - {e}')         
            else:
                # os.remove(scan.path)
                shutil.move(file_pdf_name,
                            os.path.join('.', 'dzienniki', 'error', scan.name.split('.')[0] + '.pdf'))
                error.append(file_pdf_name)            
    return krojowania, magazyn, error, brak_danych

def Zapisywanie_danych(dane, dzial, nr_zlecenia, data=False):
    if dzial == 'M':
        licznik = 0
        for each in dane:
            try:
                dziennik, dz_created = Dziennik.objects.get_or_create(
                    nr=nr_zlecenia)
                tkanina = Tkanina.objects.get(index_sap=each['index_sap'])
                index, created = WpisyMagazyn.objects.get_or_create(
                    dziennik=dziennik, tkanina=tkanina, ilosc=each['dlugosc'])
                if created:
                    licznik += 1
            except Exception as e:
                print("Niepoprawne dane")
                print(e)
                print(f"Dzial: {dzial}, Nr zlecenia: {nr_zlecenia}")
                print(each)
                return e
    if dzial == 'K':
        licznik = 0
        for each in dane:
            try:
                dziennik, dz_created = Dziennik.objects.get_or_create(
                    nr=nr_zlecenia)
                if data:
                    data = data.replace(".", "/")                    
                    dziennik.data = datetime.strptime(data, '%d/%m/%Y').date()
                    dziennik.save()
                tkanina = Tkanina.objects.get(index_sap=each['index_sap'])
                pozycja = each['lp']
                TA = each['ta']
                ilosc = each['ilosc']
                tura = each['tura']
                index, created = WpisySzwalnia.objects.get_or_create(dziennik=dziennik,
                                                                    tkanina=tkanina,
                                                                    TA=TA,
                                                                    pozycja=pozycja,
                                                                    ilosc=ilosc,
                                                                    tura=tura)
                if created:
                    licznik += 1
            except Exception as e:
                print("Niepoprawne dane")
                print(e)
                print(f"Dzial: {dzial}, Nr zlecenia: {nr_zlecenia}")
                print(each)
                ErrorLog.objects.create(error=e, funkcja=Zapisywanie_danych(), post=each)
                return e
    return licznik

def Tworzenie_zapytania(nr, root_path=os.path.join('.', 'db')):
    try:
        input_file = 'tmp/input.txt'
        if isinstance(nr, str) is False:
            nr = str(nr)
        with open(os.path.join(input_file), 'w', encoding='utf-8') as f:
            query_string = r"SELECT * FROM JOBDATA WHERE JOB_NAME='{0}'".format(
                str(nr))
            f.write(query_string)
            f.close()

        return MDB_Query(root_path)
    except Exception as e:
        print(e)

def zapytanie_paczki(comment, root_path=os.path.join('.', 'db')):
    try:
        input_file = 'tmp/input_p.txt'
        if isinstance(comment, str) is False:
            comment = str(comment)
        with open(os.path.join(input_file), 'w', encoding='utf-8') as f:
            query_string = r"SELECT * FROM JOBDATA WHERE JOB_COMMENT like '{0}%'".format(
                str(comment))
            f.write(query_string)
            f.close()
        return MDB_Query_paczka(root_path)
    except Exception as e:
        print(e)

def read_output(output_file, paczka=False):
    try:
        if os.stat(output_file).st_size == 0:
            return False
        if paczka:
            query_result = []
            with open(output_file, 'r', encoding='utf-8') as f:
                for each in f:
                    query_result.append(each)
                return query_result
        with open(output_file, 'r', encoding='utf-8') as f:
            query_result = next(f)
            return query_result.split('\t')
    except Exception as e:
        print(e)

def MDB_Query_paczka(root_path, MDB='JobListe.mdb'):
    input_file = 'tmp/input_p.txt'
    output_file = 'tmp/output_p.txt'
    cmd = r'mdb-sql -pFH -i {0} -o {1} {2}'.format(os.path.join(input_file),
                                                   os.path.join(
                                                       output_file),
                                                   os.path.join(root_path, MDB))
    call(cmd, shell=True)
    return read_output(os.path.join(output_file), paczka=True)

def MDB_Query(root_path, MDB='JobListe.mdb'):
    input_file = 'tmp/input.txt'
    output_file = 'tmp/output.txt'
    cmd = r'mdb-sql -pFH -i {0} -o {1} {2}'.format(os.path.join(input_file),
                                                   os.path.join(
                                                       output_file),
                                                   os.path.join(root_path, MDB))
    call(cmd, shell=True)
    return read_output(os.path.join(output_file))

def zmiany_w_karcie(path=os.path.join('tmp', 'szablon.xlsx'), indeks = "", tkanina = "", kolor = "", lot = "", rolka = "", dlugosc = "", data = ""):
    try:
        # Wczytanie arkusza i nazwa karty
        wb = load_workbook(path)
        ws = wb['karta']
        return True
    except Exception as e:
        print(e, 'TOMEK - poobranie workbook')
        return False

    # Wpisane w pola arksuza
    ws['C2'] = indeks
    ws['C3'] = tkanina
    ws['C4'] = kolor
    ws['C6'] = lot
    ws['C7'] = rolka
    ws['G14'] = dlugosc
    ws['F7'] = data

    #zapisanie pliku
    try:
        wb.save(os.path.join('tmp', 'wzornik_tmp.xlsx'))
    except expression as e:
        print(e, 'TOMEK - zapisanie do wzrornik_tmp.xlsx')
        return False
    print('Zapisane w {0}'.format(path))

def zamien_wzornik_na_pdf(plik=os.path.join('tmp', 'wzornik_tmp.xlsx')):
    try:
        #call(['curl --from file=@tmp/wzornik_tmp.xlsx http://uniconv/unoconv/pdf > tmp/wzornik_tmp.pdf'], shell=True)
        call(['curl --form file=@tmp/wzornik_tmp.xlsx http://uniconv:3000/unoconv/pdf > tmp/wzornik_tmp.pdf'], shell=True)
    except expression as e:
        print(e)
def typ_descr(typ):
    switcher = {
        "EDYCJA": "EDYCJA",
        "WYDANIE_MAG_WZORNIKI":"WZORNIKI",
        "WYDANIE_MAG_WYMIANKA":"WYMIANKA",
        "WPIS_MAGAZYN_WYDANIE":"WYDANIE",
        "FGK_laczone":"FGK_L",
        "FGK_laczone_poza":"FGK_LP",
        "FGK_poza":"FGK_P",
        "ODPAD":"ODPAD",
        "WPIS_MAGAZYN_DODANIE":"DODANIE",
        "WPIS_MAGAZYN_ZWROT":"ZWROT",
        "WPIS_MAGAZYN_WYCOFANIE":"WYCOFANIE",

    }
    if switcher=="":
        switcher=typ
    return switcher.get(typ, typ)


